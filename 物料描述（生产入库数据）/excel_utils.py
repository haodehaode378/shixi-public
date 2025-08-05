#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2025.07.30
# @Author : 王沁桐(3636617336@qq.com)
# @File : excel_utils.py
# @Description : 


"""Excel操作通用工具模块
包含Excel文件检查、加载、工作表获取等通用功能
"""
import os
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional


def check_file_exists(file_path: str) -> bool:
    """
    检查文件是否存在
    
    参数:
        file_path: 文件路径
        
    返回:
        存在返回True，否则返回False并打印提示
    """
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在 → {file_path}")
        return False
    return True


def load_excel_workbook(file_path: str) -> Optional[Workbook]:
    """
    加载Excel工作簿
    
    参数:
        file_path: Excel文件路径
        
    返回:
        工作簿对象，失败返回None
    """
    try:
        # data_only=True 读取公式计算后的值
        return openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"加载Excel失败: {str(e)}")
        return None


def get_excel_sheet(workbook: Workbook, sheet_name: str) -> Optional[Worksheet]:
    """
    获取指定名称的工作表
    
    参数:
        workbook: 工作簿对象
        sheet_name: 工作表名称
        
    返回:
        工作表对象，不存在返回None并打印提示
    """
    if sheet_name not in workbook.sheetnames:
        print(f"错误: 工作表「{sheet_name}」不存在！包含的表: {workbook.sheetnames}")
        return None
    return workbook[sheet_name]


def read_cell_value(sheet: Worksheet, row: int, col: int) -> str:
    """
    读取单元格值并转换为字符串
    
    参数:
        sheet: 工作表对象
        row: 行号（从1开始）
        col: 列号（从1开始）
        
    返回:
        处理后的字符串值（去除首尾空格）
    """
    cell_value = sheet.cell(row=row, column=col).value
    return str(cell_value).strip() if cell_value is not None else ""