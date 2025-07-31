import openpyxl
import pymysql
import os
import re

# ---------------------- 配置（必须修改！） ----------------------
excel_path = r"C:\Users\admin\Desktop\三江\核心产品返修率-20250611外发wqt.xlsx"  
sheet_name = "板子入库"         # Excel工作表名（必须完全一致）
CODE_COLUMN = 1                # 物料代码所在列（A列→1，B列→2...）
DESC_COLUMN = 2                # 物料描述所在列（同上）
db_config = {                  # 数据库配置
    "host": "localhost",       
    "user": "root",            
    "password": "123456",      
    "database": "三江",        
    "port": 3306               
}
target_table = "material_info" # 目标表名


# ---------------------- 数据清洗函数 ----------------------
def clean_description(desc):
    """彻底清洗描述：仅保留中文、英文、数字，去除所有特殊字符"""
    if not desc:
        return ""
    # 正则：仅保留 中文（\u4e00-\u9fa5）、英文（a-zA-Z）、数字（0-9）
    return re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', str(desc))


# ---------------------- 主逻辑 ----------------------
def main():
    # 1. 检查Excel文件存在性
    if not os.path.exists(excel_path):
        print(f"错误：Excel文件不存在 → {excel_path}")
        return

    # 2. 读取Excel文件
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in workbook.sheetnames:
            print(f"错误：工作表「{sheet_name}」不存在！Excel包含的表：{workbook.sheetnames}")
            return
        sheet = workbook[sheet_name]
        print(f"成功：加载Excel → 共 {sheet.max_row} 行数据")
        
        # 调试：打印表头，验证列索引是否正确
        header = list(sheet.rows)[0]  # 第1行是表头
        print(f"调试：表头 → 第{CODE_COLUMN}列：{header[CODE_COLUMN-1].value} | 第{DESC_COLUMN}列：{header[DESC_COLUMN-1].value}")
        if not (header[CODE_COLUMN-1].value and header[DESC_COLUMN-1].value):
            print(f"警告：表头第{CODE_COLUMN}列或第{DESC_COLUMN}列为空！请检查列索引配置。")

    except Exception as e:
        print(f"错误：读取Excel失败 → {str(e)}")
        return

    # 3. 提取并处理数据（从第2行开始）
    material_data = []
    for row_idx in range(2, sheet.max_row + 1):  # 行号从2开始（跳过表头）
        # 读取指定列的数据
        code_cell = sheet.cell(row=row_idx, column=CODE_COLUMN)
        desc_cell = sheet.cell(row=row_idx, column=DESC_COLUMN)
        
        # 强制转换为字符串，处理空值/数字类型
        raw_code = str(code_cell.value).strip() if code_cell.value is not None else ""
        raw_desc = str(desc_cell.value).strip() if desc_cell.value is not None else ""
        
        # 调试：打印原始数据
        print(f"调试：行{row_idx} → 原始代码：{raw_code} | 原始描述：{raw_desc}")
        
        # 过滤空代码
        if not raw_code:
            print(f"  跳过：行{row_idx} → 物料代码为空")
            continue
        
        # 清洗描述
        cleaned_desc = clean_description(raw_desc)
        if not cleaned_desc:
            print(f"  跳过：行{row_idx} → 描述清洗后为空（原始描述：{raw_desc}）")
            continue
        
        material_data.append((raw_code, cleaned_desc))

    # 4. 数据去重（保留第一个出现的描述）
    unique_materials = {}
    for code, desc in material_data:
        if code not in unique_materials:
            unique_materials[code] = desc
    final_data = list(unique_materials.items())
    print(f"信息：去重后 → 有效数据共 {len(final_data)} 条")

    # 5. 数据库操作
    try:
        # 连接数据库
        conn = pymysql.connect(**db_config, charset="utf8mb4")
        cursor = conn.cursor()

        # 创建表（确保支持中文）
        create_table_sql = f"""
        CREATE TABLE IF NOT EXISTS `{target_table}` (
            `material_code` VARCHAR(50) PRIMARY KEY,
            `material_desc` VARCHAR(255) NOT NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        cursor.execute(create_table_sql)
        conn.commit()
        print(f"成功：表 `{target_table}` 创建（或已存在）")

        # 插入数据（重复项自动跳过）
        if final_data:
            insert_sql = f"""
            INSERT INTO `{target_table}` (material_code, material_desc)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE material_desc = VALUES(material_desc);
            """
            cursor.executemany(insert_sql, final_data)
            conn.commit()
            print(f"成功：插入数据 → 共 {len(final_data)} 条（重复项已跳过）")
        else:
            print("警告：无有效数据可插入！可能原因：\n"
                  "  1. 列索引（CODE_COLUMN/DESC_COLUMN）配置错误\n"
                  "  2. Excel第2行及以后无有效数据\n"
                  "  3. 所有描述清洗后为空（可调整clean_description函数）")

    except pymysql.MySQLError as e:
        conn.rollback()
        print(f"错误：MySQL操作失败 → {str(e)}（建议检查：表结构、数据库权限、外键约束）")
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        print(f"错误：数据库操作失败 → {str(e)}")
    finally:
        # 关闭连接
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
        print("信息：数据库连接已关闭")


if __name__ == "__main__":
    main()
    print("=== 操作执行完毕 ===")