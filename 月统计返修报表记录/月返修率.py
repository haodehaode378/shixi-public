import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


INBOUND_PATH = r"C:\Users\admin\Desktop\入库数据分析.xlsx"   # 入库表：料号列=material_code
REPAIR_PATH  = r"C:\Users\admin\Desktop\月返修率返修统计.xlsx" # 返修表：料号列=board_code
RESULT_PATH  = r"C:\Users\admin\Desktop\月返修率计算结果.xlsx" # 结果保存路径


try:
    inbound_df = pd.read_excel(INBOUND_PATH)
    repair_df  = pd.read_excel(REPAIR_PATH)
except FileNotFoundError:
    print("文件未找到！请检查路径是否正确。")
    exit()
except Exception as e:
    print(f"读取文件失败：{e}")
    exit()

# 校验料号列和描述列是否存在
for df, name, code_col, desc_col in [
    (inbound_df, "入库表", "material_code", "material_desc"),
    (repair_df,  "返修表", "board_code",    "material_desc")
]:
    if code_col not in df.columns:
        print(f"{name} 缺少料号列：{code_col}")
        exit()
    if desc_col not in df.columns:
        print(f"{name} 缺少描述列：{desc_col}")
        exit()


months = [col for col in repair_df.columns 
          if col not in ["board_code", "material_desc"]  # 排除非月份列
          and "-" in col]  # 月份列格式：2023-01

if not months:
    print("未找到月份列！请检查列名（应为2023-01格式）。")
    exit()


merged_df = pd.merge(
    repair_df,          # 左表：返修数据（保留物料描述）
    inbound_df,         # 右表：入库数据
    left_on="board_code",   # 返修表的料号列
    right_on="material_code",# 入库表的料号列
    how="left",         # 保留返修表所有行（即使入库表无对应数据）
    suffixes=("_repair", "_inbound")  # 区分同名列（如月份列、描述列）
)


result_df = merged_df[["board_code", "material_desc_repair"]].rename(
    columns={
        "board_code": "单板料号",          # 用返修表的料号
        "material_desc_repair": "物料描述" # 用返修表的描述
    }
)

for month in months:
    repair_col = f"{month}_repair"  # 返修数列（来自返修表）
    inbound_col = f"{month}_inbound"# 入库数列（来自入库表）
    
    def calc_rate(row):
        repair = row[repair_col] if not pd.isna(row[repair_col]) else 0
        inbound = row[inbound_col] if not pd.isna(row[inbound_col]) else 0
        
        if inbound == 0:          # 入库数为0（包括无数据情况）
            return 0.00
        else:                       # 正常计算（保留2位小数）
            rate = (repair / inbound) * 100
            return round(rate, 2)
    
    result_df[month] = merged_df.apply(calc_rate, axis=1)


try:
    # 保存初始结果（使用openpyxl引擎，方便后续格式修改）
    result_df.to_excel(RESULT_PATH, index=False, engine="openpyxl")
    
    # 加载文件并设置标红格式
    wb = load_workbook(RESULT_PATH)
    ws = wb.active  # 获取活动工作表
    
    # 定义红色填充样式
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 浅红色背景
    
    # 月份列在Excel中的起始列索引（A=1, B=2, ...，前两列是"单板料号"和"物料描述"）
    start_col = 3  # 第三列开始是月份数据列
    
    # 遍历所有数据行（从第2行开始，第1行是表头）
    for row in range(2, ws.max_row + 1):
        # 遍历所有月份列
        for col in range(start_col, start_col + len(months)):
            cell = ws.cell(row=row, column=col)
            # 检查单元格值是否大于3（仅对数字有效）
            if isinstance(cell.value, (int, float)) and cell.value > 3:
                cell.fill = red_fill
    
    # 保存带格式的文件
    wb.save(RESULT_PATH)
    
    print(f"\n计算完成！结果已保存至：\n{RESULT_PATH}")
    print("说明：所有返修率以数字形式存储（如5.25代表5.25%），大于3%的单元格已标红")
except Exception as e:
    print(f"保存失败：{e}")
