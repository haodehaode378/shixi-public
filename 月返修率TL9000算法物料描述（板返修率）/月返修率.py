#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2025.07.31
# @Author : 王沁桐(3636617336@qq.com)
# @File : 月返修率.py
# @Description : 


import pandas as pd  # 用于数据处理（读取、清洗、透视表等）
import os  # 用于文件路径处理
from openpyxl import load_workbook  # 用于加载Excel文件并修改格式
from openpyxl.styles import PatternFill  # 用于设置Excel单元格背景色
from openpyxl.styles import Font
from db_utils import create_db_connection, close_db_connection  # 自定义数据库连接/关闭工具
from config import DB_CONFIG1  # 从配置文件导入数据库连接参数（主机、用户等）


def get_desktop_path():
    """
    获取Windows系统桌面路径
    
    返回:
        str: 桌面完整路径（如"C:\\Users\\用户名\\Desktop"）
    """
    # 拼接用户主目录（通过环境变量USERPROFILE获取）和"Desktop"文件夹
    return os.path.join(os.environ["USERPROFILE"], "Desktop")


def load_data():
    """
    从数据库加载入库数据和返修数据，并进行初步处理（月度汇总、时间过滤）
    
    返回:
        tuple: (stock_monthly, repair_monthly)
            stock_monthly: 入库数据月度汇总（按物料+月份统计总入库量）
            repair_monthly: 返修数据月度汇总（按单板料号+月份统计总返修量）
            若加载失败，返回(None, None)
    """
    try:
        # 建立数据库连接（通过DB_CONFIG1配置参数）
        conn = create_db_connection(**DB_CONFIG1)
        
        # 1. 读取入库数据（关联物料信息表，补充物料描述）
        stock_df = pd.read_sql(
        """
            SELECT mi.material_code, mi.material_desc, ms.date, ms.quantity 
            FROM material_stock ms  # 入库表
            JOIN material_info mi ON ms.material_code = mi.material_code  # 关联物料信息表
        """, conn)
        
        # 提取入库日期中的"年份-月份"（如"2023-05"），用于后续按月汇总
        stock_df['month'] = pd.to_datetime(stock_df['date']).dt.strftime('%Y-%m')
        
        # 按"物料编码+物料描述+月份"分组，计算每月总入库量（列名改为inbound_qty）
        stock_monthly = stock_df.groupby(
            ['material_code', 'material_desc', 'month']
        )['quantity'].sum().reset_index(name='inbound_qty')
        
        # 2. 读取返修数据
        repair_df = pd.read_sql(
            "SELECT board_code, count, year, month FROM repair_stats",  # 从返修表读取数据
            conn
        )
        # 过滤时间：只保留2023年1月及以后的数据（业务需求：关注近期返修情况）
        repair_df = repair_df[
            (repair_df['year'] > 2023) |  # 年份>2023的全部保留
            ((repair_df['year'] == 2023) & (repair_df['month'] >= 1))  # 2023年只保留1月及以后
        ]
        # 将年份和月份合并为"年份-月份"格式（如"2023-01"），确保与入库数据的月份格式一致
        repair_df['month'] = repair_df.apply(
            lambda x: f"{x['year']}-{x['month']:02d}", axis=1  # 月份补0（如1→"01"）
        )
        
        # 按"单板料号+月份"分组，计算每月总返修量（列名改为repair_qty）
        repair_monthly = repair_df.groupby(
            ['board_code', 'month']
        )['count'].sum().reset_index(name='repair_qty')
        
        # 关闭数据库连接（释放资源）
        close_db_connection(conn)
        return stock_monthly, repair_monthly  # 返回处理后的入库和返修月度数据
    
    except Exception as e:
        print(f"数据加载失败: {e}")  # 捕获异常并提示错误信息
        return None, None  # 加载失败时返回空值


def calculate_repair_rate(stock_monthly, repair_monthly):
    """
    计算单物料月度返修率和全局月度总返修率，生成透视表报表
    
    参数:
        stock_monthly: 入库数据月度汇总（load_data返回的第一个值）
        repair_monthly: 返修数据月度汇总（load_data返回的第二个值）
    
    返回:
        pd.DataFrame: 透视表报表（行：物料，列：月份，值：返修率，含全局总计行）
        若输入数据为空，返回None
    """
    # 若入库或返修数据为空，直接返回None（避免后续处理报错）
    if stock_monthly.empty or repair_monthly.empty:
        return None
    
    
    # 构建物料映射表（物料编码→物料描述，去重确保唯一对应）
    material_map = stock_monthly[['material_code', 'material_desc']].drop_duplicates()
    
    # 关联返修数据与物料信息（假设board_code=material_code，补充物料描述）
    repair_merged = pd.merge(
        repair_monthly,
        material_map,
        left_on='board_code',  # 返修数据的单板料号
        right_on='material_code',  # 物料表的物料编码
        how='left'  # 左连接：保留所有返修数据，未匹配到的物料描述为NaN
    ).drop(columns=['board_code'])  # 移除冗余的board_code列
    
    # 合并入库数据和返修数据（按物料+月份外连接，确保所有入库/返修记录都保留）
    merged_data = pd.merge(
        stock_monthly,
        repair_merged,
        on=['material_code', 'material_desc', 'month'],  # 按3个字段对齐
        how='outer'  # 外连接：两边数据都保留，无匹配的字段用NaN填充
    )
    
    # 填充空值：入库量/返修量为空时视为0（无入库/无返修）
    merged_data[['inbound_qty', 'repair_qty']] = merged_data[['inbound_qty', 'repair_qty']].fillna(0)
    
    
    def calc_monthly_rate(row):
        """
        计算单物料当月返修率（返修量÷入库量×100%，保留2位小数）
        """
        inbound, repair = row['inbound_qty'], row['repair_qty']
        # 若入库量为0或返修量为0，返修率视为0（避免除0错误或无意义数据）
        return 0.00 if (inbound == 0 or repair == 0) else round((repair / inbound) * 100, 2)
    
    # 应用函数，生成单物料月度返修率列
    merged_data['monthly_rate(%)'] = merged_data.apply(calc_monthly_rate, axis=1)
    
    
    # 按月份分组，统计当月所有物料的总入库量和总返修量
    monthly_global = merged_data.groupby('month').agg({
        'inbound_qty': 'sum',  # 当月总入库量（所有物料之和）
        'repair_qty': 'sum'    # 当月总返修量（所有物料之和）
    }).reset_index()
    # 计算全局月度返修率（总返修÷总入库×100%，保留2位小数）
    monthly_global['global_monthly_rate(%)'] = monthly_global.apply(
        lambda x: round((x['repair_qty'] / x['inbound_qty'] * 100) if x['inbound_qty'] != 0 else 0, 2),
        axis=1
    )
    
    
    pivot = merged_data.pivot_table(
        index=['material_code', 'material_desc'],  # 行：物料编码+物料描述
        columns='month',  # 列：月份（如"2023-01"）
        values='monthly_rate(%)',  # 单元格值：单物料当月返修率
        aggfunc='sum',  # 聚合方式：求和（同一物料同月唯一，sum不影响结果）
        fill_value=0  # 空值填充为0（无数据的月份返修率视为0）
    ).reset_index()  # 重置索引，将行索引转为普通列
    
    
    if not pivot.empty:
        # 提取所有月份列（排除物料编码和描述列）
        month_cols = [col for col in pivot.columns if col not in ['material_code', 'material_desc']]
        # 按时间顺序排序月份（如"2023-01"→"2023-02"）
        sorted_months = sorted(month_cols, key=lambda x: pd.to_datetime(x))
        # 按排序后的月份重新排列列顺序
        pivot = pivot[['material_code', 'material_desc'] + sorted_months]
        
        
        total_row = {'material_code': '', 'material_desc': '当月全局总计'}  # 总计行标识
        # 为每个月份列填充全局返修率
        for month in sorted_months:
            # 从全局月度表中匹配对应月份的总返修率
            rate = monthly_global[monthly_global['month'] == month]['global_monthly_rate(%)'].values[0] if not monthly_global.empty else 0.0
            total_row[month] = rate
        
        # 将总计行添加到透视表末尾
        pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)
    
    return pivot  # 返回最终的透视表报表


def export_report(report_df):
    """
    将返修率报表导出到桌面Excel，并设置红色背景（返修率>3%的单元格）
    
    参数:
        report_df: 待导出的透视表报表（calculate_repair_rate的返回值）
    """
    # 若报表为空，提示并退出
    if report_df is None or report_df.empty:
        print("无有效数据，无法导出")
        return
    
    # 定义导出路径：桌面+固定文件名
    file_path = os.path.join(get_desktop_path(), "月返修率百分比统计.xlsx")
    
    try:
        # 第一步：将报表数据导出到Excel（不保留索引）
        report_df.to_excel(file_path, index=False, engine='openpyxl')
        
        # 第二步：加载Excel并设置格式（大于3%的单元格标红色背景）
        wb = load_workbook(file_path)  # 重新加载刚导出的Excel文件
        ws = wb.active  # 获取当前活跃的工作表（默认第一个表）
        
        # 定义红色背景样式（浅红色，RGB编码FFC7CE，solid填充）
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="FF0000")
        
        # 确定需要处理的单元格范围：
        month_start_col = 3  # 月份数据从第3列开始（前2列是物料编码和描述）
        month_end_col = ws.max_column  # 月份列的最后一列
        data_end_row = ws.max_row  # 数据的最后一行（含总计行）

        # 遍历所有数据行（跳过表头行，从第2行开始）
        for row in range(2, data_end_row + 1):
            # 遍历所有月份列（从第3列到最后一列）
            for col in range(month_start_col, month_end_col + 1):
                cell = ws.cell(row=row, column=col)  # 获取当前单元格
                try:
                    value = float(cell.value)  # 将单元格值转为数字
                    if value > 1:  # 判断是否大于3%
                        cell.font = red_font  # 设置为红色字体
                except (ValueError, TypeError):
                    continue  # 非数值类型（如空值、字符串）不处理
        
        # 遍历所有数据单元格（跳过表头行，从第2行开始）
        for row in range(2, data_end_row + 1):  # 行：从第2行到最后一行
            for col in range(month_start_col, month_end_col + 1):  # 列：从月份列到最后一列
                cell = ws.cell(row=row, column=col)  # 获取当前单元格
                # 尝试将单元格值转为数字，判断是否大于3%
                try:
                    value = float(cell.value)  # 转换为浮点数
                    if value > 3:  # 若大于3%，设置红色背景
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    # 非数值类型（如空值、字符串）不处理，避免报错
                    continue
        
        # 保存格式修改
        wb.save(file_path)
        print(f"报表已保存至：{file_path}（大于3%的数据已设置红色背景）")
    
    except Exception as e:
        print(f"导出失败：{e}")  # 捕获导出过程中的异常并提示


def main():
    """
    程序主入口：协调数据加载→返修率计算→报表导出全流程
    """
    # 1. 加载入库和返修数据
    stock_data, repair_data = load_data()
    # 若数据加载失败，退出程序
    if stock_data is None or repair_data is None:
        return
    
    # 2. 计算返修率并生成报表
    report = calculate_repair_rate(stock_data, repair_data)
    
    # 3. 导出报表到桌面
    export_report(report)


# 当脚本直接运行时，执行主函数
if __name__ == "__main__":
    main()