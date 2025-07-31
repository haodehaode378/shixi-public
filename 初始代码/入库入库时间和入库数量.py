import openpyxl
import pymysql

# 配置项（请根据实际情况修改）
EXCEL_FILE = r"C:\Users\admin\Desktop\三江\核心产品返修率-20250611外发wqt.xlsx"
SHEET_NAME = "板子入库"
DB_CONFIG = {
    'host': 'localhost',    
    'user': 'root',         
    'password': '123456',   # 替换为实际密码
    'database': '三江'      # 替换为实际数据库名
}
target_table = "material_stock"  # 目标表名

def main():
    # 1. 加载Excel文件（读取公式计算后的值）
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        if SHEET_NAME not in workbook.sheetnames:
            print(f"错误：工作表 '{SHEET_NAME}' 不存在")
            return
        sheet = workbook[SHEET_NAME]
        print("Excel文件加载成功")
    except Exception as e:
        print(f"打开Excel失败：{e}")
        return

    # 2. 连接数据库
    try:
        conn = pymysql.connect(**DB_CONFIG, charset="utf8mb4")
        cursor = conn.cursor()
        print("数据库连接成功")
    except Exception as e:
        print(f"数据库连接失败：{e}")
        workbook.close()
        return

    try:
        # 3. 创建表（包含所需字段）
        create_table_sql = f"""
        CREATE TABLE IF NOT EXISTS `{target_table}` (
            `material_code` VARCHAR(50),  -- 允许为空（取消NOT NULL约束）
            `seq` VARCHAR(20),            -- 允许为空
            `date` DATE,                  -- 允许为空
            `quantity` FLOAT,             -- 允许为空
            PRIMARY KEY (`material_code`, `seq`, `date`)  -- 联合主键（空值可能导致插入失败）
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        cursor.execute(create_table_sql)
        conn.commit()
        print(f"表 `{target_table}` 检查/创建成功")

        # 4. 遍历数据并插入数据库（不跳过任何记录）
        success_count = 0
        fail_count = 0
        
        # 保持原有的行范围（2到182行）
        for row in range(2, 183):
            # 提取物料代码（A列，column=1）
            material_code = sheet.cell(row=row, column=1).value
            # 提取序号（C列，column=3）
            seq = sheet.cell(row=row, column=3).value
            
            # 遍历时间列（H→AJ，列8到36）
            for col in range(8, 37):
                # 获取数量值
                quantity = sheet.cell(row=row, column=col).value
                # 获取日期（表头行第1行）
                date_str = sheet.cell(row=1, column=col).value
                
                # 不跳过任何数据，即使有空值也尝试插入
                try:
                    sql = f"""
                    INSERT INTO `{target_table}` 
                    (material_code, seq, date, quantity)
                    VALUES (%s, %s, %s, %s)
                    """
                    cursor.execute(sql, (material_code, seq, date_str, quantity))
                    conn.commit()
                    success_count += 1
                except Exception as e:
                    print(f"插入失败：{e}，数据：{material_code}, {seq}, {date_str}, {quantity}")
                    conn.rollback()
                    fail_count += 1

        print(f"\n数据导入完成！成功：{success_count}条，失败：{fail_count}条")

    except Exception as e:
        print(f"执行过程出错：{e}")
    finally:
        # 5. 关闭所有资源
        cursor.close()
        conn.close()
        workbook.close()
        print("资源已释放")

if __name__ == "__main__":
    main()